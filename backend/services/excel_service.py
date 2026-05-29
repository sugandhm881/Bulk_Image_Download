import re
from io import BytesIO
from openpyxl import load_workbook

VALID_EXTENSIONS = (
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".bmp",
    ".webp",
)


def extract_urls_from_text(text):

    if text is None:
        return []

    text = str(text)

    pattern = r'https?://[^\s,"\']+'

    matches = re.findall(pattern, text)

    valid_urls = []

    for url in matches:
        if any(ext in url.lower() for ext in VALID_EXTENSIONS):
            valid_urls.append(url)

    return valid_urls


def read_rows(file_bytes):
    """Read the first worksheet into (headers, list-of-dict-rows) using openpyxl.

    Avoids pandas/numpy so the function stays small enough for serverless.
    """

    wb = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], []

    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]

    data = []

    for r in rows[1:]:
        row = {}
        for i, h in enumerate(headers):
            row[h] = r[i] if i < len(r) else None
        data.append(row)

    return headers, data
